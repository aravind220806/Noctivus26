import io
import pytest
from openpyxl import load_workbook

from app.db.sqlite_db import sqlite_db
from app.events import EVENT_CATALOG
from app.services.admin_access_service import ADMIN_TABS
from app.services.export_service import export_attendance_to_excel
from app.services.registration_service import serialize_registration
from app.routes.admin_routes import (
    extract_event_members_with_attendance,
    find_registration_flexible,
    format_registration_for_attendance,
)


def test_admin_tabs_includes_attendance():
    assert "Attendance" in ADMIN_TABS


@pytest.mark.asyncio
async def test_find_registration_flexible_and_members_extraction():
    await sqlite_db.init()

    sample_reg = {
        "registrationId": "NOC26-TEST99",
        "qrToken": "test_qr_token_12345",
        "qrHash": "dummy_hash_value",
        "paymentStatus": "confirmed",
        "participant": {
            "name": "ALICE SMITH",
            "email": "alice@example.com",
            "phone": "9876543210",
            "college": "Velammal Engineering College",
            "department": "CSE",
            "year": "3",
            "rollNo": "22CS101",
        },
        "eventRegistrations": [
            {
                "eventId": "prompt-heist",
                "eventName": "Prompt Heist",
                "category": "tech",
                "teamSize": 2,
                "teamMembers": [
                    {"name": "BOB JONES", "rollNo": "22CS102"}
                ],
                "attendance": {
                    "attended": True,
                    "markedAt": "2026-09-26T10:30:00Z",
                    "markedBy": "coordinator@noctivus26.com",
                    "members": [
                        {"name": "ALICE SMITH", "rollNo": "22CS101", "role": "Team Leader", "isLeader": True, "present": True},
                        {"name": "BOB JONES", "rollNo": "22CS102", "role": "Team Member", "isLeader": False, "present": False},
                    ]
                }
            }
        ],
        "checkedIn": True,
    }

    await sqlite_db.upsert("registrations", "NOC26-TEST99", sample_reg)

    # 1. Test flexible lookup by ID
    found_by_id = await find_registration_flexible("NOC26-TEST99")
    assert found_by_id is not None
    assert found_by_id["registrationId"] == "NOC26-TEST99"

    # 2. Test flexible lookup by QR token / URL
    found_by_url = await find_registration_flexible("https://noctivus26.com/p/test_qr_token_12345")
    assert found_by_url is not None
    assert found_by_url["registrationId"] == "NOC26-TEST99"

    # 3. Test flexible lookup by email
    found_by_email = await find_registration_flexible("alice@example.com")
    assert found_by_email is not None
    assert found_by_email["registrationId"] == "NOC26-TEST99"

    # 4. Test members extraction
    members = extract_event_members_with_attendance(sample_reg, "prompt-heist")
    assert len(members) == 2
    leader = next((m for m in members if m["isLeader"]), None)
    assert leader is not None
    assert leader["name"] == "ALICE SMITH"
    assert leader["present"] is True

    teammate = next((m for m in members if not m["isLeader"]), None)
    assert teammate is not None
    assert teammate["name"] == "BOB JONES"
    assert teammate["present"] is False

    # 5. Test formatted representation
    formatted = format_registration_for_attendance(sample_reg)
    assert "eventAttendanceList" in formatted
    ev_info = formatted["eventAttendanceList"][0]
    assert ev_info["eventId"] == "prompt-heist"
    assert ev_info["presentCount"] == 1
    assert ev_info["totalCount"] == 2
    assert ev_info["isPartial"] is True

    # Cleanup
    await sqlite_db.delete("registrations", "NOC26-TEST99")


def test_export_attendance_to_excel():
    sample_registrations = [
        {
            "registrationId": "NOC26-WIN01",
            "paymentStatus": "confirmed",
            "checkedIn": True,
            "participant": {
                "name": "SARAH CONNOR",
                "email": "sarah@cyber.io",
                "phone": "9998887776",
                "college": "Tech Institute of Chennai",
                "department": "IT",
                "year": "4",
                "rollNo": "IT401",
            },
            "eventRegistrations": [
                {
                    "eventId": "ctf",
                    "eventName": "NULL CORE 2.0 CTF",
                    "category": "tech",
                    "teamSize": 3,
                    "teamMembers": [
                        {"name": "JOHN CONNOR", "rollNo": "IT402"},
                        {"name": "KYLE REESE", "rollNo": "IT403"},
                    ],
                    "attendance": {
                        "attended": True,
                        "markedAt": "2026-09-26T10:15:00Z",
                        "markedBy": "ctf_coord@noctivus26.com",
                        "members": [
                            {"name": "SARAH CONNOR", "rollNo": "IT401", "role": "Team Leader", "isLeader": True, "present": True},
                            {"name": "JOHN CONNOR", "rollNo": "IT402", "role": "Team Member", "isLeader": False, "present": True},
                            {"name": "KYLE REESE", "rollNo": "IT403", "role": "Team Member", "isLeader": False, "present": False},
                        ]
                    }
                }
            ]
        },
        {
            "registrationId": "NOC26-SOLO1",
            "paymentStatus": "confirmed",
            "checkedIn": True,
            "participant": {
                "name": "NEO ANDERSON",
                "email": "neo@matrix.io",
                "phone": "9112223334",
                "college": "Velammal Engineering College",
                "department": "CSE",
                "year": "3",
                "rollNo": "CS301",
            },
            "eventRegistrations": [
                {
                    "eventId": "bug-hunt",
                    "eventName": "Bug Hunt",
                    "category": "tech",
                    "teamSize": 1,
                    "teamMembers": [],
                    "attendance": {
                        "attended": True,
                        "markedAt": "2026-09-26T11:00:00Z",
                        "markedBy": "bughunt_coord@noctivus26.com",
                        "members": [
                            {"name": "NEO ANDERSON", "rollNo": "CS301", "role": "Team Leader", "isLeader": True, "present": True},
                        ]
                    }
                }
            ]
        }
    ]

    excel_bytes = export_attendance_to_excel(EVENT_CATALOG[:5], sample_registrations)
    assert isinstance(excel_bytes, bytes)
    assert len(excel_bytes) > 0

    wb = load_workbook(io.BytesIO(excel_bytes))
    sheet_names = wb.sheetnames

    # Check Master Attendance Summary
    assert "Attendance Summary" in sheet_names
    summary_ws = wb["Attendance Summary"]
    assert summary_ws.cell(row=1, column=1).value == "S.No"
    assert summary_ws.cell(row=1, column=3).value == "Event Name"

    # Check E-Cert Master List sheet
    assert "E-Cert Master List" in sheet_names
    ecert_ws = wb["E-Cert Master List"]
    assert ecert_ws.cell(row=1, column=3).value == "Member Full Name"

    # Verify present members in E-Cert master sheet
    present_names = [ecert_ws.cell(row=r, column=3).value for r in range(2, ecert_ws.max_row + 1)]
    assert "SARAH CONNOR" in present_names
    assert "JOHN CONNOR" in present_names
    assert "NEO ANDERSON" in present_names
    assert "KYLE REESE" not in present_names # KYLE was absent

    # Check that individual event sheets exist
    assert any("CTF" in name or "NULL" in name for name in sheet_names)
    assert any("Bug Hunt" in name for name in sheet_names)
