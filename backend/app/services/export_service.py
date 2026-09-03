import csv
import io
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def _safe_csv_value(value):
    if not isinstance(value, str):
        return value
    if value.lstrip().startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def _get_abstract(registration: dict, event_id: str | None = None) -> str:
    if not registration:
        return ""
    event_regs = registration.get("eventRegistrations") or []
    if event_id:
        for ev in event_regs:
            if ev.get("eventId") == event_id and ev.get("abstract"):
                return ev["abstract"]
    for ev in event_regs:
        if ev.get("abstract"):
            return ev["abstract"]
    participant = registration.get("participant") or {}
    return (
        registration.get("abstract")
        or registration.get("igniteTopic")
        or participant.get("abstract")
        or participant.get("igniteTopic")
        or ""
    )


def registrations_to_csv(registrations: list[dict], sponsor_safe: bool = False, event_id: str | None = None) -> str:
    output = io.StringIO()
    writer = csv.writer(output)
    if sponsor_safe:
        writer.writerow(["Name", "College", "Events", "Abstract / Topic"])
    else:
        writer.writerow([
            "Registration ID",
            "Name",
            "Email",
            "Phone",
            "College",
            "Food",
            "Events",
            "Abstract / Topic",
            "Status",
            "UTR",
            "Expected Amount",
            "Claimed Amount",
            "Checked In",
            "Checked In At",
            "Submitted At",
            "Verified At",
        ])
    for registration in registrations:
        participant = registration.get("participant") or {}
        abstract_val = _get_abstract(registration, event_id)
        values = [
            participant.get("name"),
            participant.get("college"),
            "; ".join(event.get("eventName", "") for event in registration.get("eventRegistrations", [])),
            abstract_val,
        ] if sponsor_safe else [
            registration.get("registrationId"),
            participant.get("name"),
            participant.get("email"),
            participant.get("phone"),
            participant.get("college"),
            participant.get("foodPreference"),
            "; ".join(event.get("eventName", "") for event in registration.get("eventRegistrations", [])),
            abstract_val,
            registration.get("paymentStatus"),
            registration.get("utrNumber"),
            registration.get("expectedAmount"),
            registration.get("claimedAmount"),
            registration.get("checkedIn", False),
            registration.get("checkedInAt"),
            registration.get("paymentSubmittedAt"),
            registration.get("verifiedAt"),
        ]
        writer.writerow(_safe_csv_value(value) for value in values)
    return output.getvalue()


def export_scheduler_to_excel(events: list[dict], slots: list[dict], registrations: list[dict]) -> bytes:
    wb = Workbook()
    
    # Define styles
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    morning_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    afternoon_fill = PatternFill(start_color="E0F2FE", end_color="E0F2FE", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

    reg_by_id = {
        (r.get("registrationId") or r.get("member_id")): r
        for r in registrations
    }

    # ================= SHEET 1: Master Slot Schedule =================
    ws1 = wb.active
    ws1.title = "Master Event Slots"

    headers1 = [
        "Event Name",
        "Category",
        "Window",
        "Date",
        "Start Time",
        "End Time",
        "Slot Timing",
        "Capacity",
        "Assigned Count",
        "Available",
        "Assigned Member IDs",
        "Assigned Member Names",
    ]
    ws1.append(headers1)

    for col_num, header in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    events_map = {e["id"]: e for e in events}
    
    # Sort slots: by event name, then window (morning first), then start_time
    sorted_slots = sorted(
        slots,
        key=lambda s: (
            events_map.get(s.get("event_id"), {}).get("name", s.get("event_id", "")),
            0 if str(s.get("window")).lower() == "morning" else 1,
            s.get("start_time", ""),
        ),
    )

    for row_idx, slot in enumerate(sorted_slots, 2):
        ev = events_map.get(slot.get("event_id"), {})
        assigned_ids = slot.get("assigned_member_ids") or []
        assigned_count = len(assigned_ids)
        capacity = slot.get("capacity") or 30
        available = max(0, capacity - assigned_count)
        
        member_names = []
        for mid in assigned_ids:
            reg = reg_by_id.get(mid)
            if reg:
                pname = (reg.get("participant") or {}).get("name", mid)
                member_names.append(f"{pname} ({mid})")
            else:
                member_names.append(mid)

        window_label = "Morning" if str(slot.get("window")).lower() == "morning" else "Afternoon"
        slot_timing = f"{slot.get('start_time', '')} - {slot.get('end_time', '')}"

        row_data = [
            ev.get("name", slot.get("event_id")),
            ev.get("category", "tech"),
            window_label,
            slot.get("date", "2026-09-26"),
            slot.get("start_time", ""),
            slot.get("end_time", ""),
            slot_timing,
            capacity,
            assigned_count,
            available,
            ", ".join(assigned_ids) if assigned_ids else "None",
            "; ".join(member_names) if member_names else "None",
        ]
        ws1.append(row_data)

        # Apply cell borders and subtle window colors
        for col_idx in range(1, len(row_data) + 1):
            cell = ws1.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            if col_idx == 3: # Window column
                cell.fill = morning_fill if window_label == "Morning" else afternoon_fill
                cell.alignment = Alignment(horizontal="center")
            elif col_idx in (4, 5, 6, 7, 8, 9, 10):
                cell.alignment = Alignment(horizontal="center")

    # ================= SHEET 2: Event Summary =================
    ws2 = wb.create_sheet(title="Event Summary")
    headers2 = [
        "Event ID",
        "Event Name",
        "Category",
        "Duration (Mins)",
        "Total Registrations",
        "Total Slots",
        "Morning Slots",
        "Afternoon Slots",
        "Total Capacity",
        "Total Assigned",
        "Utilization %",
    ]
    ws2.append(headers2)
    for col_num in range(1, len(headers2) + 1):
        cell = ws2.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    slots_by_ev = {}
    for s in slots:
        eid = s.get("event_id")
        if eid not in slots_by_ev:
            slots_by_ev[eid] = []
        slots_by_ev[eid].append(s)

    for row_idx, ev in enumerate(events, 2):
        eid = ev["id"]
        ev_slots = slots_by_ev.get(eid, [])
        morning_count = sum(1 for s in ev_slots if str(s.get("window")).lower() == "morning")
        afternoon_count = sum(1 for s in ev_slots if str(s.get("window")).lower() == "afternoon")
        total_capacity = sum(s.get("capacity", 30) for s in ev_slots)
        total_assigned = sum(len(s.get("assigned_member_ids", [])) for s in ev_slots)
        utilization = f"{(total_assigned / total_capacity * 100):.1f}%" if total_capacity > 0 else "0.0%"

        row_data = [
            eid,
            ev.get("name", eid),
            ev.get("category", "tech"),
            ev.get("duration_minutes", 90),
            ev.get("total_registrations", 0),
            len(ev_slots),
            morning_count,
            afternoon_count,
            total_capacity,
            total_assigned,
            utilization,
        ]
        ws2.append(row_data)
        for col_idx in range(1, len(row_data) + 1):
            cell = ws2.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            if col_idx in (4, 5, 6, 7, 8, 9, 10, 11):
                cell.alignment = Alignment(horizontal="center")

    # ================= SHEET 3: Member Allocations =================
    ws3 = wb.create_sheet(title="Member Allocations")
    headers3 = [
        "Registration ID",
        "Member Name",
        "Email",
        "College",
        "Registered Events",
        "Abstract / Topic",
        "Assigned Slot IDs",
        "Slot Details (Window & Timing)",
    ]
    ws3.append(headers3)
    for col_num in range(1, len(headers3) + 1):
        cell = ws3.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    slots_map = {s["id"]: s for s in slots}
    
    confirmed_regs = [r for r in registrations if r.get("paymentStatus") == "confirmed"]
    for row_idx, reg in enumerate(confirmed_regs, 2):
        participant = reg.get("participant") or {}
        assigned_slot_ids = reg.get("assigned_slots") or []
        
        slot_descriptions = []
        for sid in assigned_slot_ids:
            sl = slots_map.get(sid)
            if sl:
                ev_name = events_map.get(sl.get("event_id"), {}).get("name", sl.get("event_id"))
                win = "Morning" if sl.get("window") == "morning" else "Afternoon"
                slot_descriptions.append(f"{ev_name}: {win} ({sl.get('start_time')} - {sl.get('end_time')})")
            else:
                slot_descriptions.append(sid)

        event_names = [e.get("eventName") or e.get("eventId") for e in reg.get("eventRegistrations", [])]

        row_data = [
            reg.get("registrationId") or reg.get("member_id"),
            participant.get("name", ""),
            participant.get("email", ""),
            participant.get("college", ""),
            ", ".join(event_names),
            _get_abstract(reg),
            ", ".join(assigned_slot_ids) if assigned_slot_ids else "Unassigned",
            "; ".join(slot_descriptions) if slot_descriptions else "Unassigned",
        ]
        ws3.append(row_data)
        for col_idx in range(1, len(row_data) + 1):
            cell = ws3.cell(row=row_idx, column=col_idx)
            cell.border = thin_border

    # Auto-adjust column widths on all sheets
    for sheet in (ws1, ws2, ws3):
        for col in sheet.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

def _sanitize_sheet_title(name: str) -> str:
    r"""Excel sheet names max length is 31 and cannot contain : \ / ? * [ ]"""
    clean = re.sub(r'[:\\/?*\[\]]', '', str(name or 'Event')).strip()
    return clean[:31] if clean else 'Event'


def export_attendance_to_excel(events: list[dict], registrations: list[dict]) -> bytes:
    wb = Workbook()

    primary_header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    primary_header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    present_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    present_font = Font(name="Calibri", size=10, bold=True, color="166534")

    absent_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    absent_font = Font(name="Calibri", size=10, color="991B1B")

    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    def get_event_members(reg: dict, target_event_id: str) -> list[dict]:
        p = reg.get("participant") or {}
        event_regs = reg.get("eventRegistrations") or []
        ev_item = next((e for e in event_regs if e.get("eventId") == target_event_id), None)
        if not ev_item:
            return []

        att_data = ev_item.get("attendance") or (reg.get("attendance") or {}).get(target_event_id) or {}
        members_att = {
            (m.get("name") or "").strip().upper(): m.get("present", False)
            for m in att_data.get("members", [])
            if isinstance(m, dict)
        }

        marked_at = att_data.get("markedAt") or ""
        marked_by = att_data.get("markedBy") or ""

        members_list = []

        leader_name = (p.get("name") or "").strip().upper()
        leader_present = members_att.get(leader_name, att_data.get("present", False) if "members" not in att_data else False)

        members_list.append({
            "registrationId": reg.get("registrationId") or reg.get("member_id", ""),
            "name": leader_name,
            "role": "Team Leader",
            "rollNo": p.get("rollNo") or p.get("collegeId") or "",
            "college": p.get("college", ""),
            "department": p.get("department", ""),
            "year": p.get("year", ""),
            "email": p.get("email", ""),
            "phone": p.get("phone", ""),
            "paymentStatus": reg.get("paymentStatus", "pending"),
            "checkedIn": "Yes" if reg.get("checkedIn") else "No",
            "present": leader_present,
            "markedAt": marked_at,
            "markedBy": marked_by,
        })

        raw_team_members = ev_item.get("teamMembers") or []
        for tm in raw_team_members:
            if not isinstance(tm, dict):
                continue
            tm_name = (tm.get("name") or "").strip().upper()
            if not tm_name:
                continue
            tm_present = members_att.get(tm_name, False)
            members_list.append({
                "registrationId": reg.get("registrationId") or reg.get("member_id", ""),
                "name": tm_name,
                "role": "Team Member",
                "rollNo": tm.get("rollNo", ""),
                "college": p.get("college", ""),
                "department": p.get("department", ""),
                "year": p.get("year", ""),
                "email": p.get("email", ""),
                "phone": p.get("phone", ""),
                "paymentStatus": reg.get("paymentStatus", "pending"),
                "checkedIn": "Yes" if reg.get("checkedIn") else "No",
                "present": tm_present,
                "markedAt": marked_at,
                "markedBy": marked_by,
            })

        return members_list

    # ================= SHEET 1: Master Attendance Summary =================
    ws_summary = wb.active
    ws_summary.title = "Attendance Summary"

    summary_headers = [
        "S.No",
        "Event ID",
        "Event Name",
        "Category",
        "Venue",
        "Date & Time",
        "Teams Registered",
        "Total Members",
        "Present Members",
        "Absent Members",
        "Attendance %",
    ]
    ws_summary.append(summary_headers)

    for col_num in range(1, len(summary_headers) + 1):
        cell = ws_summary.cell(row=1, column=col_num)
        cell.fill = primary_header_fill
        cell.font = primary_header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    all_present_members_master = []

    for s_idx, ev in enumerate(events, 1):
        eid = ev.get("id")
        ename = ev.get("name", eid)
        ecat = ev.get("category", "tech")
        evenue = ev.get("venue", "Campus")
        edatetime = f"{ev.get('date', '2026-09-26')} {ev.get('time', '10:00 AM')}"

        event_regs = [
            r for r in registrations
            if any(e.get("eventId") == eid for e in r.get("eventRegistrations", []))
        ]

        total_teams = len(event_regs)
        all_event_members = []
        for r in event_regs:
            m_list = get_event_members(r, eid)
            all_event_members.extend(m_list)
            for m in m_list:
                if m.get("present"):
                    all_present_members_master.append({
                        **m,
                        "eventId": eid,
                        "eventName": ename,
                        "category": ecat,
                    })

        total_members_count = len(all_event_members)
        present_count = sum(1 for m in all_event_members if m.get("present"))
        absent_count = max(0, total_members_count - present_count)
        rate_str = f"{(present_count / total_members_count * 100):.1f}%" if total_members_count > 0 else "0.0%"

        ws_summary.append([
            s_idx,
            eid,
            ename,
            ecat,
            evenue,
            edatetime,
            total_teams,
            total_members_count,
            present_count,
            absent_count,
            rate_str,
        ])

        summary_row_num = s_idx + 1
        for col_idx in range(1, len(summary_headers) + 1):
            cell = ws_summary.cell(row=summary_row_num, column=col_idx)
            cell.border = thin_border
            if col_idx in (1, 7, 8, 9, 10, 11):
                cell.alignment = Alignment(horizontal="center")

        # ================= EVENT SPECIFIC SHEET =================
        sheet_name = _sanitize_sheet_title(ename)
        if sheet_name in wb.sheetnames:
            sheet_name = _sanitize_sheet_title(f"{ename[:26]}-{s_idx}")

        ws_event = wb.create_sheet(title=sheet_name)

        event_sheet_headers = [
            "S.No",
            "Registration ID",
            "Member Name (E-Certificate Name)",
            "Role",
            "Roll No / ID",
            "College",
            "Department",
            "Year",
            "Email",
            "Phone",
            "Payment Status",
            "Gate Check-In",
            "Event Attendance",
            "E-Certificate Eligible",
            "Attendance Marked At",
            "Marked By",
        ]
        ws_event.append(event_sheet_headers)

        for col_num in range(1, len(event_sheet_headers) + 1):
            cell = ws_event.cell(row=1, column=col_num)
            cell.fill = primary_header_fill
            cell.font = primary_header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        sorted_event_members = sorted(
            all_event_members,
            key=lambda x: (0 if x.get("present") else 1, x.get("registrationId", ""), 0 if x.get("role") == "Team Leader" else 1)
        )

        for m_idx, m in enumerate(sorted_event_members, 1):
            is_present = m.get("present", False)
            is_confirmed = m.get("paymentStatus") == "confirmed"
            is_cert_eligible = "YES" if (is_present and is_confirmed) else "NO"
            att_status_text = "PRESENT" if is_present else "ABSENT"

            row_data = [
                m_idx,
                m.get("registrationId", ""),
                m.get("name", ""),
                m.get("role", ""),
                m.get("rollNo", ""),
                m.get("college", ""),
                m.get("department", ""),
                m.get("year", ""),
                m.get("email", ""),
                m.get("phone", ""),
                m.get("paymentStatus", "").capitalize(),
                m.get("checkedIn", "No"),
                att_status_text,
                is_cert_eligible,
                m.get("markedAt", ""),
                m.get("markedBy", ""),
            ]
            ws_event.append(row_data)

            curr_row = m_idx + 1
            for col_idx in range(1, len(row_data) + 1):
                cell = ws_event.cell(row=curr_row, column=col_idx)
                cell.border = thin_border
                if col_idx in (1, 4, 8, 11, 12, 13, 14):
                    cell.alignment = Alignment(horizontal="center")
                if col_idx in (13, 14):
                    if is_present and is_cert_eligible == "YES":
                        cell.fill = present_fill
                        cell.font = present_font
                    else:
                        cell.fill = absent_fill
                        cell.font = absent_font

    # ================= SHEET: All Present Members (E-Certificate Master List) =================
    ws_master = wb.create_sheet(title="E-Cert Master List")
    master_headers = [
        "S.No",
        "Registration ID",
        "Member Full Name",
        "Role",
        "Roll No / ID",
        "College",
        "Department",
        "Year",
        "Email",
        "Phone",
        "Event Name",
        "Category",
        "Attendance Status",
        "E-Certificate Status",
        "Marked At",
        "Marked By",
    ]
    ws_master.append(master_headers)
    for col_num in range(1, len(master_headers) + 1):
        cell = ws_master.cell(row=1, column=col_num)
        cell.fill = primary_header_fill
        cell.font = primary_header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for p_idx, pm in enumerate(all_present_members_master, 1):
        master_row = [
            p_idx,
            pm.get("registrationId", ""),
            pm.get("name", ""),
            pm.get("role", ""),
            pm.get("rollNo", ""),
            pm.get("college", ""),
            pm.get("department", ""),
            pm.get("year", ""),
            pm.get("email", ""),
            pm.get("phone", ""),
            pm.get("eventName", ""),
            pm.get("category", ""),
            "PRESENT",
            "APPROVED FOR E-CERTIFICATE",
            pm.get("markedAt", ""),
            pm.get("markedBy", ""),
        ]
        ws_master.append(master_row)
        curr_row = p_idx + 1
        for col_idx in range(1, len(master_row) + 1):
            cell = ws_master.cell(row=curr_row, column=col_idx)
            cell.border = thin_border
            if col_idx in (1, 4, 8, 12, 13, 14):
                cell.alignment = Alignment(horizontal="center")
            if col_idx in (13, 14):
                cell.fill = present_fill
                cell.font = present_font

    # Auto-adjust column widths on ALL sheets
    for sheet in wb.worksheets:
        sheet.views.sheetView[0].showGridLines = True
        for col in sheet.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 40)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def export_full_live_backup_excel(events: list[dict], registrations: list[dict], slots: list[dict]) -> bytes:
    """
    Generates a master live-backup Excel workbook containing:
    1. Registered (All registrations)
    2. Verified (Confirmed payment registrations)
    3. Check-In List (Gate check-in participants)
    4. Master Event Slots (Scheduler slots)
    5. Scheduler Summary (Event capacities & allocations)
    6. Member Slot Allocations (Participant slot timings)
    7+. Dedicated attendance sheets for every event
    """
    wb = Workbook()

    # Styling
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    badge_green_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    badge_green_font = Font(name="Calibri", size=10, bold=True, color="166534")
    badge_red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    badge_red_font = Font(name="Calibri", size=10, color="991B1B")
    morning_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    afternoon_fill = PatternFill(start_color="E0F2FE", end_color="E0F2FE", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    events_map = {e["id"]: e for e in events}
    slots_map = {s["id"]: s for s in slots}

    # ================= 1. SHEET: Registered =================
    ws_reg = wb.active
    ws_reg.title = "Registered"
    reg_headers = [
        "S.No",
        "Registration ID",
        "Participant Name",
        "Email",
        "Phone",
        "College",
        "Department",
        "Year",
        "Food Preference",
        "Registered Events",
        "Abstract / Topic",
        "Payment Status",
        "UTR Number",
        "Expected Amount",
        "Claimed Amount",
        "Gate Check-In",
        "Checked In At",
        "Submitted At",
        "Verified At",
    ]
    ws_reg.append(reg_headers)
    for col_num in range(1, len(reg_headers) + 1):
        cell = ws_reg.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r_idx, reg in enumerate(registrations, 1):
        p = reg.get("participant") or {}
        event_names = [e.get("eventName") or e.get("eventId") for e in reg.get("eventRegistrations", [])]
        is_checked_in = bool(reg.get("checkedIn"))
        p_status = (reg.get("paymentStatus") or "pending").capitalize()

        row_data = [
            r_idx,
            reg.get("registrationId") or reg.get("member_id", ""),
            p.get("name", ""),
            p.get("email", ""),
            p.get("phone", ""),
            p.get("college", ""),
            p.get("department", ""),
            p.get("year", ""),
            p.get("foodPreference", ""),
            "; ".join(event_names),
            _get_abstract(reg),
            p_status,
            reg.get("utrNumber", ""),
            reg.get("expectedAmount", 0),
            reg.get("claimedAmount", 0),
            "YES" if is_checked_in else "NO",
            reg.get("checkedInAt", ""),
            reg.get("paymentSubmittedAt", "") or reg.get("createdAt", ""),
            reg.get("verifiedAt", ""),
        ]
        ws_reg.append(row_data)
        curr_row = r_idx + 1
        for col_idx in range(1, len(row_data) + 1):
            cell = ws_reg.cell(row=curr_row, column=col_idx)
            cell.border = thin_border
            if col_idx in (1, 8, 9, 12, 14, 15, 16):
                cell.alignment = Alignment(horizontal="center")
            if col_idx == 12:  # Payment status
                if p_status.lower() == "confirmed":
                    cell.fill = badge_green_fill
                    cell.font = badge_green_font
            elif col_idx == 16:  # Gate Check-In
                if is_checked_in:
                    cell.fill = badge_green_fill
                    cell.font = badge_green_font

    # ================= 2. SHEET: Verified =================
    ws_ver = wb.create_sheet(title="Verified")
    ver_headers = [
        "S.No",
        "Registration ID",
        "Participant Name",
        "Email",
        "Phone",
        "College",
        "Department",
        "Year",
        "Food",
        "Registered Events",
        "UTR Number",
        "Verified Amount",
        "Verified At",
        "Gate Check-In",
    ]
    ws_ver.append(ver_headers)
    for col_num in range(1, len(ver_headers) + 1):
        cell = ws_ver.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    verified_regs = [r for r in registrations if (r.get("paymentStatus") or "").lower() == "confirmed"]
    for v_idx, reg in enumerate(verified_regs, 1):
        p = reg.get("participant") or {}
        event_names = [e.get("eventName") or e.get("eventId") for e in reg.get("eventRegistrations", [])]
        is_checked_in = bool(reg.get("checkedIn"))

        row_data = [
            v_idx,
            reg.get("registrationId") or reg.get("member_id", ""),
            p.get("name", ""),
            p.get("email", ""),
            p.get("phone", ""),
            p.get("college", ""),
            p.get("department", ""),
            p.get("year", ""),
            p.get("foodPreference", ""),
            "; ".join(event_names),
            reg.get("utrNumber", ""),
            reg.get("expectedAmount", 0),
            reg.get("verifiedAt", ""),
            "YES" if is_checked_in else "NO",
        ]
        ws_ver.append(row_data)
        curr_row = v_idx + 1
        for col_idx in range(1, len(row_data) + 1):
            cell = ws_ver.cell(row=curr_row, column=col_idx)
            cell.border = thin_border
            if col_idx in (1, 8, 9, 12, 14):
                cell.alignment = Alignment(horizontal="center")
            if col_idx == 14 and is_checked_in:
                cell.fill = badge_green_fill
                cell.font = badge_green_font

    # ================= 3. SHEET: Check-In List =================
    ws_chk = wb.create_sheet(title="Check-In List")
    chk_headers = [
        "S.No",
        "Registration ID",
        "Participant Name",
        "Email",
        "Phone",
        "College",
        "Department",
        "Food",
        "Registered Events",
        "Payment Status",
        "Checked In At",
        "Checked In By",
    ]
    ws_chk.append(chk_headers)
    for col_num in range(1, len(chk_headers) + 1):
        cell = ws_chk.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    checked_in_regs = [r for r in registrations if bool(r.get("checkedIn"))]
    for c_idx, reg in enumerate(checked_in_regs, 1):
        p = reg.get("participant") or {}
        event_names = [e.get("eventName") or e.get("eventId") for e in reg.get("eventRegistrations", [])]

        row_data = [
            c_idx,
            reg.get("registrationId") or reg.get("member_id", ""),
            p.get("name", ""),
            p.get("email", ""),
            p.get("phone", ""),
            p.get("college", ""),
            p.get("department", ""),
            p.get("foodPreference", ""),
            "; ".join(event_names),
            (reg.get("paymentStatus") or "pending").capitalize(),
            reg.get("checkedInAt", ""),
            reg.get("checkedInBy", "Gate Desk"),
        ]
        ws_chk.append(row_data)
        curr_row = c_idx + 1
        for col_idx in range(1, len(row_data) + 1):
            cell = ws_chk.cell(row=curr_row, column=col_idx)
            cell.border = thin_border
            if col_idx in (1, 8, 10):
                cell.alignment = Alignment(horizontal="center")

    # ================= 4. SHEET: Master Event Slots =================
    ws_slots = wb.create_sheet(title="Master Event Slots")
    slots_headers = [
        "Event Name",
        "Category",
        "Window",
        "Date",
        "Start Time",
        "End Time",
        "Slot Timing",
        "Capacity",
        "Assigned Count",
        "Available",
        "Assigned Member IDs",
        "Assigned Member Names",
    ]
    ws_slots.append(slots_headers)
    for col_num in range(1, len(slots_headers) + 1):
        cell = ws_slots.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    reg_by_id = {(r.get("registrationId") or r.get("member_id")): r for r in registrations}
    sorted_slots = sorted(
        slots,
        key=lambda s: (
            events_map.get(s.get("event_id"), {}).get("name", s.get("event_id", "")),
            0 if str(s.get("window")).lower() == "morning" else 1,
            s.get("start_time", ""),
        ),
    )
    for s_idx, slot in enumerate(sorted_slots, 2):
        ev = events_map.get(slot.get("event_id"), {})
        assigned_ids = slot.get("assigned_member_ids") or []
        assigned_count = len(assigned_ids)
        capacity = slot.get("capacity") or 30
        available = max(0, capacity - assigned_count)
        member_names = []
        for mid in assigned_ids:
            reg = reg_by_id.get(mid)
            if reg:
                pname = (reg.get("participant") or {}).get("name", mid)
                member_names.append(f"{pname} ({mid})")
            else:
                member_names.append(mid)

        window_label = "Morning" if str(slot.get("window")).lower() == "morning" else "Afternoon"
        slot_timing = f"{slot.get('start_time', '')} - {slot.get('end_time', '')}"

        row_data = [
            ev.get("name", slot.get("event_id")),
            ev.get("category", "tech"),
            window_label,
            slot.get("date", "2026-09-26"),
            slot.get("start_time", ""),
            slot.get("end_time", ""),
            slot_timing,
            capacity,
            assigned_count,
            available,
            ", ".join(assigned_ids) if assigned_ids else "None",
            "; ".join(member_names) if member_names else "None",
        ]
        ws_slots.append(row_data)
        for col_idx in range(1, len(row_data) + 1):
            cell = ws_slots.cell(row=s_idx, column=col_idx)
            cell.border = thin_border
            if col_idx == 3:
                cell.fill = morning_fill if window_label == "Morning" else afternoon_fill
                cell.alignment = Alignment(horizontal="center")
            elif col_idx in (4, 5, 6, 7, 8, 9, 10):
                cell.alignment = Alignment(horizontal="center")

    # ================= 5. SHEET: Scheduler Event Summary =================
    ws_sched_sum = wb.create_sheet(title="Scheduler Summary")
    sched_sum_headers = [
        "Event ID",
        "Event Name",
        "Category",
        "Duration (Mins)",
        "Total Registrations",
        "Total Slots",
        "Morning Slots",
        "Afternoon Slots",
        "Total Capacity",
        "Total Assigned",
        "Utilization %",
    ]
    ws_sched_sum.append(sched_sum_headers)
    for col_num in range(1, len(sched_sum_headers) + 1):
        cell = ws_sched_sum.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    slots_by_ev = {}
    for s in slots:
        eid = s.get("event_id")
        if eid not in slots_by_ev:
            slots_by_ev[eid] = []
        slots_by_ev[eid].append(s)

    for row_idx, ev in enumerate(events, 2):
        eid = ev["id"]
        ev_slots = slots_by_ev.get(eid, [])
        morning_count = sum(1 for s in ev_slots if str(s.get("window")).lower() == "morning")
        afternoon_count = sum(1 for s in ev_slots if str(s.get("window")).lower() == "afternoon")
        total_capacity = sum(s.get("capacity", 30) for s in ev_slots)
        total_assigned = sum(len(s.get("assigned_member_ids", [])) for s in ev_slots)
        utilization = f"{(total_assigned / total_capacity * 100):.1f}%" if total_capacity > 0 else "0.0%"

        row_data = [
            eid,
            ev.get("name", eid),
            ev.get("category", "tech"),
            ev.get("duration_minutes", 90),
            ev.get("total_registrations", 0),
            len(ev_slots),
            morning_count,
            afternoon_count,
            total_capacity,
            total_assigned,
            utilization,
        ]
        ws_sched_sum.append(row_data)
        for col_idx in range(1, len(row_data) + 1):
            cell = ws_sched_sum.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            if col_idx in (4, 5, 6, 7, 8, 9, 10, 11):
                cell.alignment = Alignment(horizontal="center")

    # ================= 6. SHEET: Member Slot Allocations =================
    ws_alloc = wb.create_sheet(title="Member Allocations")
    alloc_headers = [
        "Registration ID",
        "Member Name",
        "Email",
        "College",
        "Registered Events",
        "Abstract / Topic",
        "Assigned Slot IDs",
        "Slot Details (Window & Timing)",
    ]
    ws_alloc.append(alloc_headers)
    for col_num in range(1, len(alloc_headers) + 1):
        cell = ws_alloc.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, reg in enumerate(verified_regs, 2):
        participant = reg.get("participant") or {}
        assigned_slot_ids = reg.get("assigned_slots") or []
        slot_descriptions = []
        for sid in assigned_slot_ids:
            sl = slots_map.get(sid)
            if sl:
                ev_name = events_map.get(sl.get("event_id"), {}).get("name", sl.get("event_id"))
                win = "Morning" if sl.get("window") == "morning" else "Afternoon"
                slot_descriptions.append(f"{ev_name}: {win} ({sl.get('start_time')} - {sl.get('end_time')})")
            else:
                slot_descriptions.append(sid)

        event_names = [e.get("eventName") or e.get("eventId") for e in reg.get("eventRegistrations", [])]
        row_data = [
            reg.get("registrationId") or reg.get("member_id"),
            participant.get("name", ""),
            participant.get("email", ""),
            participant.get("college", ""),
            ", ".join(event_names),
            _get_abstract(reg),
            ", ".join(assigned_slot_ids) if assigned_slot_ids else "Unassigned",
            "; ".join(slot_descriptions) if slot_descriptions else "Unassigned",
        ]
        ws_alloc.append(row_data)
        for col_idx in range(1, len(row_data) + 1):
            cell = ws_alloc.cell(row=row_idx, column=col_idx)
            cell.border = thin_border

    # ================= 7+. SHEETS: Per-Event Attendance Sheets =================
    def _extract_members(reg: dict, target_event_id: str) -> list[dict]:
        p = reg.get("participant") or {}
        event_regs = reg.get("eventRegistrations") or []
        ev_item = next((e for e in event_regs if e.get("eventId") == target_event_id), None)
        if not ev_item:
            return []
        att_data = ev_item.get("attendance") or (reg.get("attendance") or {}).get(target_event_id) or {}
        members_att = {
            (m.get("name") or "").strip().upper(): m.get("present", False)
            for m in att_data.get("members", [])
            if isinstance(m, dict)
        }
        marked_at = att_data.get("markedAt") or ""
        marked_by = att_data.get("markedBy") or ""
        members_list = []

        leader_name = (p.get("name") or "").strip().upper()
        leader_present = members_att.get(leader_name, att_data.get("present", False) if "members" not in att_data else False)
        members_list.append({
            "registrationId": reg.get("registrationId") or reg.get("member_id", ""),
            "name": leader_name,
            "role": "Team Leader",
            "rollNo": p.get("rollNo") or p.get("collegeId") or "",
            "college": p.get("college", ""),
            "department": p.get("department", ""),
            "year": p.get("year", ""),
            "email": p.get("email", ""),
            "phone": p.get("phone", ""),
            "paymentStatus": reg.get("paymentStatus", "pending"),
            "checkedIn": "Yes" if reg.get("checkedIn") else "No",
            "present": leader_present,
            "markedAt": marked_at,
            "markedBy": marked_by,
        })

        for tm in (ev_item.get("teamMembers") or []):
            if not isinstance(tm, dict):
                continue
            tm_name = (tm.get("name") or "").strip().upper()
            if not tm_name:
                continue
            tm_present = members_att.get(tm_name, False)
            members_list.append({
                "registrationId": reg.get("registrationId") or reg.get("member_id", ""),
                "name": tm_name,
                "role": "Team Member",
                "rollNo": tm.get("rollNo", ""),
                "college": p.get("college", ""),
                "department": p.get("department", ""),
                "year": p.get("year", ""),
                "email": p.get("email", ""),
                "phone": p.get("phone", ""),
                "paymentStatus": reg.get("paymentStatus", "pending"),
                "checkedIn": "Yes" if reg.get("checkedIn") else "No",
                "present": tm_present,
                "markedAt": marked_at,
                "markedBy": marked_by,
            })
        return members_list

    for s_idx, ev in enumerate(events, 1):
        eid = ev.get("id")
        ename = ev.get("name", eid)
        clean_sheet_name = _sanitize_sheet_title(f"Att - {ename}")
        if clean_sheet_name in wb.sheetnames:
            clean_sheet_name = _sanitize_sheet_title(f"Att - {ename[:22]}-{s_idx}")

        ws_ev = wb.create_sheet(title=clean_sheet_name)
        ev_headers = [
            "S.No",
            "Registration ID",
            "Member Name (E-Certificate Name)",
            "Role",
            "Roll No / ID",
            "College",
            "Department",
            "Year",
            "Email",
            "Phone",
            "Payment Status",
            "Gate Check-In",
            "Event Attendance",
            "E-Certificate Eligible",
            "Attendance Marked At",
            "Marked By",
        ]
        ws_ev.append(ev_headers)
        for col_num in range(1, len(ev_headers) + 1):
            cell = ws_ev.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        event_regs = [
            r for r in registrations
            if any(e.get("eventId") == eid for e in r.get("eventRegistrations", []))
        ]
        all_event_members = []
        for r in event_regs:
            all_event_members.extend(_extract_members(r, eid))

        sorted_event_members = sorted(
            all_event_members,
            key=lambda x: (0 if x.get("present") else 1, x.get("registrationId", ""), 0 if x.get("role") == "Team Leader" else 1)
        )

        for m_idx, m in enumerate(sorted_event_members, 1):
            is_present = m.get("present", False)
            is_confirmed = (m.get("paymentStatus") or "").lower() == "confirmed"
            is_cert_eligible = "YES" if (is_present and is_confirmed) else "NO"
            att_status_text = "PRESENT" if is_present else "ABSENT"

            row_data = [
                m_idx,
                m.get("registrationId", ""),
                m.get("name", ""),
                m.get("role", ""),
                m.get("rollNo", ""),
                m.get("college", ""),
                m.get("department", ""),
                m.get("year", ""),
                m.get("email", ""),
                m.get("phone", ""),
                (m.get("paymentStatus") or "").capitalize(),
                m.get("checkedIn", "No"),
                att_status_text,
                is_cert_eligible,
                m.get("markedAt", ""),
                m.get("markedBy", ""),
            ]
            ws_ev.append(row_data)
            curr_row = m_idx + 1
            for col_idx in range(1, len(row_data) + 1):
                cell = ws_ev.cell(row=curr_row, column=col_idx)
                cell.border = thin_border
                if col_idx in (1, 4, 8, 11, 12, 13, 14):
                    cell.alignment = Alignment(horizontal="center")
                if col_idx in (13, 14):
                    if is_present and is_cert_eligible == "YES":
                        cell.fill = badge_green_fill
                        cell.font = badge_green_font
                    elif not is_present:
                        cell.fill = badge_red_fill
                        cell.font = badge_red_font

    # ================= 8+. SHEETS: Per-Event Slot Allocation Sheets =================
    for s_idx, ev in enumerate(events, 1):
        eid = ev.get("id")
        ename = ev.get("name", eid)
        clean_slot_sheet_name = _sanitize_sheet_title(f"Slots - {ename}")
        if clean_slot_sheet_name in wb.sheetnames:
            clean_slot_sheet_name = _sanitize_sheet_title(f"Slots - {ename[:20]}-{s_idx}")

        ws_ev_slot = wb.create_sheet(title=clean_slot_sheet_name)
        slot_headers = [
            "S.No",
            "Slot Window",
            "Timing",
            "Date",
            "Capacity",
            "Assigned",
            "Available",
            "Assigned Reg ID",
            "Member Name",
            "College",
            "Department",
            "Year",
            "Email",
            "Phone",
            "Payment Status",
            "Gate Check-In",
        ]
        ws_ev_slot.append(slot_headers)
        for col_num in range(1, len(slot_headers) + 1):
            cell = ws_ev_slot.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ev_slots = [s for s in slots if s.get("event_id") == eid]
        ev_slots_sorted = sorted(
            ev_slots,
            key=lambda s: (0 if str(s.get("window")).lower() == "morning" else 1, s.get("start_time", ""))
        )

        row_counter = 1
        for sl in ev_slots_sorted:
            assigned_ids = sl.get("assigned_member_ids") or []
            assigned_count = len(assigned_ids)
            capacity = sl.get("capacity") or 30
            available = max(0, capacity - assigned_count)
            window_label = "Morning" if str(sl.get("window")).lower() == "morning" else "Afternoon"
            timing_label = f"{sl.get('start_time', '')} - {sl.get('end_time', '')}"
            date_label = sl.get("date", "2026-09-26")

            if not assigned_ids:
                row_data = [
                    row_counter,
                    window_label,
                    timing_label,
                    date_label,
                    capacity,
                    0,
                    available,
                    "None",
                    "Unassigned / Open Slot",
                    "-",
                    "-",
                    "-",
                    "-",
                    "-",
                    "-",
                    "-",
                ]
                ws_ev_slot.append(row_data)
                curr_row = row_counter + 1
                for col_idx in range(1, len(row_data) + 1):
                    cell = ws_ev_slot.cell(row=curr_row, column=col_idx)
                    cell.border = thin_border
                    if col_idx in (1, 2, 4, 5, 6, 7):
                        cell.alignment = Alignment(horizontal="center")
                    if col_idx == 2:
                        cell.fill = morning_fill if window_label == "Morning" else afternoon_fill
                row_counter += 1
            else:
                for mid in assigned_ids:
                    reg = reg_by_id.get(mid)
                    p = (reg.get("participant") or {}) if reg else {}
                    is_chk = reg and reg.get("checkedIn")
                    row_data = [
                        row_counter,
                        window_label,
                        timing_label,
                        date_label,
                        capacity,
                        assigned_count,
                        available,
                        mid,
                        p.get("name", mid),
                        p.get("college", ""),
                        p.get("department", ""),
                        p.get("year", ""),
                        p.get("email", ""),
                        p.get("phone", ""),
                        (reg.get("paymentStatus") or "").capitalize() if reg else "",
                        "YES" if is_chk else "NO",
                    ]
                    ws_ev_slot.append(row_data)
                    curr_row = row_counter + 1
                    for col_idx in range(1, len(row_data) + 1):
                        cell = ws_ev_slot.cell(row=curr_row, column=col_idx)
                        cell.border = thin_border
                        if col_idx in (1, 2, 4, 5, 6, 7, 15, 16):
                            cell.alignment = Alignment(horizontal="center")
                        if col_idx == 2:
                            cell.fill = morning_fill if window_label == "Morning" else afternoon_fill
                        if col_idx == 16 and is_chk:
                            cell.fill = badge_green_fill
                            cell.font = badge_green_font
                    row_counter += 1

    # Auto-adjust column widths on ALL sheets
    for sheet in wb.worksheets:
        sheet.views.sheetView[0].showGridLines = True
        for col in sheet.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 40)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
